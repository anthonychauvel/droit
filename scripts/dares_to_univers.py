#!/usr/bin/env python3
"""
Construit idcc_list_complet.txt (l'univers de référence des IDCC) à partir du
fichier officiel DARES "Suivi historique des conventions collectives".

POURQUOI CE SCRIPT ?
L'univers était jusqu'ici produit par list_all_ccn.py, qui interroge Légifrance
en mode exploratoire. Cet appel échoue (HTTP 500 côté Légifrance), donc
idcc_list_complet.txt reste vide -- et comme le workflow protège la récupération
par « if [ -s idcc_list_complet.txt ] », plus aucune nouvelle convention n'était
récupérée.

DARES publie la liste officielle et exhaustive des IDCC, avec pour chacun :
  - IDCCactif : 1 si la convention est toujours en vigueur
  - NouvIDCC  : l'IDCC repreneur lorsqu'elle a fusionné
C'est une source plus fiable et plus stable qu'un scraping d'API, et elle ne
dépend pas de la disponibilité de Légifrance.

CE QUE LE SCRIPT ÉCRIT
  - idcc_list_complet.txt : les IDCC ACTIFS uniquement, un par ligne.
    C'est l'univers auquel detect_nouveaux.py compare la liste suivie.
  - idcc_fusions.csv (optionnel) : ancien_idcc,nouvel_idcc pour les conventions
    fusionnées. Utile pour prévenir un utilisateur que sa convention a été
    absorbée par une autre.

MISE À JOUR DU FICHIER SOURCE
DARES republie ce fichier plusieurs fois par an. Il suffit de remplacer le .xlsx
dans le dépôt : le script relit tout à chaque run.

Usage :
    python3 dares_to_univers.py --xlsx ccn/Dares_Suivi_Historique_convention_collective_Juin2026.xlsx \\
        --out idcc_list_complet.txt --fusions idcc_fusions.csv
"""
import argparse
import os
import sys


def lire_dares(chemin):
    """Renvoie (actifs, fusions) depuis la feuille 'Conventions de branche'.

    actifs  : set d'entiers (IDCC en vigueur)
    fusions : dict ancien_idcc -> nouvel_idcc (uniquement quand renseigné)
    """
    try:
        import openpyxl
    except ImportError:
        print("openpyxl est requis : pip install openpyxl", file=sys.stderr)
        raise

    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    nom = "Conventions de branche"
    if nom not in wb.sheetnames:
        raise SystemExit(f"Feuille '{nom}' absente de {chemin} — "
                         f"feuilles trouvées : {wb.sheetnames}")

    lignes = wb[nom].iter_rows(values_only=True)
    entete = next(lignes, None)
    if not entete:
        raise SystemExit(f"{chemin} : feuille vide.")

    # On repère les colonnes par leur intitulé plutôt que par leur position :
    # DARES peut réordonner les colonnes d'une publication à l'autre.
    col = {}
    for i, h in enumerate(entete):
        cle = str(h or "").strip().lower()
        if cle == "idcc":            col["idcc"] = i
        elif cle == "idccactif":     col["actif"] = i
        elif cle == "nouvidcc":      col["nouv"] = i
    manquantes = {"idcc", "actif"} - set(col)
    if manquantes:
        raise SystemExit(f"Colonnes introuvables dans {chemin} : {manquantes}. "
                         f"En-tête lu : {entete}")

    actifs, fusions = set(), {}
    for r in lignes:
        brut = r[col["idcc"]] if col["idcc"] < len(r) else None
        if brut is None:
            continue
        texte = str(brut).strip()
        if not texte.isdigit():
            continue
        idcc = int(texte)          # "00016" -> 16, format attendu par le fonds

        if str(r[col["actif"]] if col["actif"] < len(r) else "").strip() == "1":
            actifs.add(idcc)
        elif "nouv" in col and col["nouv"] < len(r):
            nv = str(r[col["nouv"]] or "").strip()
            if nv.isdigit() and int(nv) != idcc:
                fusions[idcc] = int(nv)

    return actifs, fusions


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Fichier DARES (.xlsx)")
    ap.add_argument("--out", default="idcc_list_complet.txt")
    ap.add_argument("--fusions", help="CSV optionnel ancien,nouveau")
    ap.add_argument("--min-attendu", type=int, default=300,
                    help="Garde-fou : en dessous de ce nombre d'IDCC actifs, on "
                         "considère la lecture ratée et on n'écrit rien.")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"{args.xlsx} introuvable — univers inchangé.", file=sys.stderr)
        return 1

    actifs, fusions = lire_dares(args.xlsx)

    # Même précaution que dans detect_nouveaux.py : ne jamais écraser l'univers
    # par un résultat manifestement incomplet. Un fichier tronqué ferait passer
    # tout le fonds pour disparu au run suivant.
    if len(actifs) < args.min_attendu:
        print(f"Seulement {len(actifs)} IDCC actifs lus (seuil : {args.min_attendu}) — "
              f"lecture douteuse, {args.out} laissé inchangé.", file=sys.stderr)
        return 1

    with open(args.out, "w", encoding="utf-8") as f:
        f.write("\n".join(str(i) for i in sorted(actifs)) + "\n")
    print(f"{args.out} : {len(actifs)} IDCC actifs écrits.")

    if args.fusions:
        with open(args.fusions, "w", encoding="utf-8") as f:
            f.write("ancien_idcc,nouvel_idcc\n")
            for a in sorted(fusions):
                f.write(f"{a},{fusions[a]}\n")
        print(f"{args.fusions} : {len(fusions)} fusions enregistrées.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
